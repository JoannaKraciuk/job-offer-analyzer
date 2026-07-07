from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path
import re
import unicodedata
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from job_offer_analyzer.config.columns import col
from job_offer_analyzer.models import (
    AvailabilityRefreshRow,
    AvailabilityRefreshSummary,
    CvMatchRefreshRow,
    CvMatchRefreshSummary,
    HistoryRecord,
    OfferRecord,
    SalaryInfo,
    SalaryRefreshRow,
    SalaryRefreshSummary,
    UNKNOWN_VALUE,
)
from job_offer_analyzer.offer_availability import check_offer_availability


OFFERS_SHEET = "Oferty"
HISTORY_SHEET = "Historia_Sprawdzeń"
QUESTIONS_SHEET = "Pytania_Formularzy"
CV_MATCH_SHEET = "Analiza_CV"
LEGACY_CV_MATCH_SHEET = "Dopasowanie_CV"
DASHBOARD_SHEET = "Dashboard"

OFFER_ID_PATTERN = re.compile(r"^JOB-(\d+)$")
POLISH_TRANSLATION = str.maketrans(
    "ąćęłńóśźżĄĆĘŁŃÓŚŹŻ",
    "acelnoszzACELNOSZZ",
)
HEADER_ALIASES = {
    col("match_score"): [col("match_score_legacy")],
    col("priority_code"): [col("priority_code_legacy")],
    col("contract_type"): ["Forma"],
    col("rate_expectations"): [
        col("rate_expectations_legacy"),
        col("rate_expectations_usd_legacy"),
    ],
}
SHEET_ALIASES = {
    CV_MATCH_SHEET: [LEGACY_CV_MATCH_SHEET],
}
OFFER_HEADER_RENAMES = {
    "Dostepnosc": col("availability"),
    "Forma": col("contract_type"),
    col("rate_expectations_legacy"): col("rate_expectations"),
    col("match_score_legacy"): col("match_score"),
    "Must-have skrot": col("must_have_summary"),
    "Nice-to-have skrot": col("nice_to_have_summary"),
    "Nastepny krok": col("next_step"),
    "Zrodlo": col("source"),
    col("cv_match"): col("cv_match"),
    col("priority"): col("priority"),
}
HISTORY_HEADER_RENAMES = {
    "Zrodlo": col("source"),
}
CV_MATCH_HEADER_RENAMES = {
    col("in_cv_legacy"): col("in_profile"),
    col("cv_evidence_legacy"): col("evidence_skill"),
    col("cv_improvement_legacy"): col("missing_to_learn"),
    col("note"): col("comment"),
}
DASHBOARD_TEXT_RENAMES = {
    "Baza ofert pracy - dashboard": "Baza ofert pracy - dashboard",
    "Wartosc": "Wartość",
    "Najblizsze dzialania": "Najbliższe działania",
    "Dostepne": "Dostępne",
    "Nastepny krok": col("next_step"),
    "Wymagaja sprawdzenia >14 dni": "Wymagają sprawdzenia >14 dni",
}
VALUE_RENAMES = {
    "TBD": UNKNOWN_VALUE,
    "Unknown": "Nieznany",
    "Dostepna": "Dostępna",
    "Zamknieta": "Zamknięta",
    "Sredni": "Średni",
    "Pobrac tresc oferty i wykonac analize pod CV": (
        "Pobrać treść oferty i wykonać analizę pod CV"
    ),
    "Porownac z CV i przygotowac odpowiedzi do formularza": (
        "Porównać z CV i przygotować odpowiedzi do formularza"
    ),
}
ACTIVE_DASHBOARD_EXCLUDED_STATUSES = {"Aplikowano", "Odrzucona"}
ACTIVE_DASHBOARD_EXCLUDED_AVAILABILITY = {"Zamknięta", "Zamknieta"}
PRIORITY_CODE_TO_LABEL = {
    "HIGH": "Wysoki",
    "MEDIUM": "Średni",
    "LOW": "Niski",
}
PRIORITY_LABEL_TO_CODE = {
    "wysoki": "HIGH",
    "sredni": "MEDIUM",
    "niski": "LOW",
}
TECHNOLOGY_ALIASES = {
    "Playwright": ["playwright"],
    "Cypress": ["cypress"],
    "Selenium": ["selenium"],
    "Python": ["python", "pytest"],
    "JavaScript": ["javascript", "js"],
    "TypeScript": ["typescript", "ts"],
    "Java": ["java"],
    "C#": ["c#", "c sharp", "csharp"],
    "API": ["api testing", "api tests", "rest api", "postman", "api"],
    "SQL": ["sql"],
    "WCAG": ["wcag", "accessibility", "a11y"],
    "Allure": ["allure"],
    "Git": ["git", "github", "gitlab", "bitbucket"],
    "CI/CD": ["ci/cd", "cicd", "github actions", "jenkins"],
    "Docker": ["docker"],
    "Jira": ["jira", "confluence"],
    "TestRail": ["testrail", "test rail"],
    "Robot Framework": ["robot framework"],
    "Cucumber": ["cucumber", "gherkin", "bdd"],
    "Manual testing": ["manual testing", "manual tests"],
    "Test automation": ["test automation", "automated tests", "automation testing"],
    "Regression testing": ["regression testing", "regression tests"],
    "Performance testing": ["performance testing", "load testing", "jmeter"],
    "Security testing": ["security testing", "penetration testing"],
    "Mobile testing": ["mobile testing", "android", "ios"],
    "Agile": ["agile", "scrum", "kanban"],
    "AI coding assistants": ["ai coding", "ai assistant", "copilot", "chatgpt"],
}
SALARY_HEADERS = [
    col("salary_source"),
    col("currency"),
    col("salary_min"),
    col("salary_max"),
    col("salary_period"),
    col("tax_type"),
    col("exchange_rate"),
    col("exchange_rate_date"),
    col("pln_min_monthly"),
    col("pln_max_monthly"),
    col("pln_min_hourly"),
    col("pln_max_hourly"),
    col("conversion_assumptions"),
]
OFFER_TRACKING_HEADERS = [
    col("match_score"),
    col("priority"),
    col("priority_code"),
    col("technologies"),
    col("portal"),
    col("last_action"),
]
OFFER_MATCH_HEADERS = [col("missing_skills")]
CV_MATCH_HEADERS = [col("url")]
AUTO_MATCH_CATEGORY = "Automatyczne dopasowanie"
TABLE_NAMES = {
    OFFERS_SHEET: "OfertyTable",
    CV_MATCH_SHEET: "DopasowanieTable",
    LEGACY_CV_MATCH_SHEET: "DopasowanieTable",
    QUESTIONS_SHEET: "PytaniaTable",
    HISTORY_SHEET: "HistoriaTable",
}


def append_offer_to_workbook(workbook_path: Path, offer: OfferRecord) -> str:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")
    _ensure_workbook_writable(workbook_path)

    workbook = load_workbook(workbook_path)
    _normalize_workbook_sheets(workbook)
    offers_sheet = _get_sheet(workbook, OFFERS_SHEET)
    history_sheet = _get_sheet(workbook, HISTORY_SHEET)
    questions_sheet = _get_sheet(workbook, QUESTIONS_SHEET)
    dashboard_sheet = _get_sheet(workbook, DASHBOARD_SHEET)
    _normalize_offer_headers(offers_sheet)
    _normalize_sheet_headers(history_sheet, HISTORY_HEADER_RENAMES)
    _ensure_sheet_headers(offers_sheet, SALARY_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_TRACKING_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_MATCH_HEADERS)
    _normalize_sheet_values(offers_sheet)
    _normalize_sheet_values(history_sheet)
    _normalize_history_actions(history_sheet)
    _normalize_dashboard_sheet(dashboard_sheet)
    _backfill_offer_tracking_columns(offers_sheet, history_sheet=history_sheet)

    duplicate_offer_id = _find_offer_id_by_link(offers_sheet, offer.link)
    if duplicate_offer_id:
        _append_history_row(
            history_sheet,
            _history_duplicate_offer(duplicate_offer_id, offer),
        )
        _refresh_dashboard_actions(dashboard_sheet, offers_sheet)
        _resize_known_table(offers_sheet)
        _resize_known_table(questions_sheet)
        _resize_known_table(history_sheet)
        workbook.save(workbook_path)
        return duplicate_offer_id

    offer_id = _next_offer_id(offers_sheet)
    _append_offer_row(offers_sheet, offer_id, offer)
    _append_history_row(history_sheet, _history_from_offer(offer_id, offer))
    _append_question_placeholder_row(questions_sheet, offer_id, offer)
    _refresh_dashboard_actions(dashboard_sheet, offers_sheet)
    _resize_known_table(offers_sheet)
    _resize_known_table(questions_sheet)
    _resize_known_table(history_sheet)

    workbook.save(workbook_path)
    return offer_id


def refresh_offer_availability_in_workbook(
    workbook_path: Path,
) -> AvailabilityRefreshSummary:
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")
    _ensure_workbook_writable(workbook_path)

    workbook = load_workbook(workbook_path)
    _normalize_workbook_sheets(workbook)
    offers_sheet = _get_sheet(workbook, OFFERS_SHEET)
    history_sheet = _get_sheet(workbook, HISTORY_SHEET)
    dashboard_sheet = _get_sheet(workbook, DASHBOARD_SHEET)
    _normalize_offer_headers(offers_sheet)
    _normalize_sheet_headers(history_sheet, HISTORY_HEADER_RENAMES)
    _ensure_sheet_headers(offers_sheet, SALARY_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_TRACKING_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_MATCH_HEADERS)
    _normalize_sheet_values(offers_sheet)
    _normalize_sheet_values(history_sheet)
    _normalize_history_actions(history_sheet)
    _normalize_dashboard_sheet(dashboard_sheet)
    _backfill_offer_tracking_columns(offers_sheet, history_sheet=history_sheet)

    headers = _header_positions(offers_sheet)
    results: list[AvailabilityRefreshRow] = []
    today = date.today()

    for row in range(2, offers_sheet.max_row + 1):
        offer_id = _cell_by_header(offers_sheet, row, headers, col("offer_id"))
        link = _cell_by_header(offers_sheet, row, headers, col("url"))
        if not offer_id or not link:
            continue

        company = _cell_by_header(offers_sheet, row, headers, col("company")) or ""
        title = _cell_by_header(offers_sheet, row, headers, col("job_title")) or ""
        previous_availability = (
            _cell_by_header(offers_sheet, row, headers, col("availability")) or UNKNOWN_VALUE
        )
        availability_result = check_offer_availability(str(link))

        _set_cell_by_header(
            offers_sheet, row, headers, col("availability"), availability_result.availability
        )
        _set_cell_by_header(offers_sheet, row, headers, col("last_checked"), today)
        _set_cell_by_header(offers_sheet, row, headers, col("days_since_check"), 0)
        _set_cell_by_header(offers_sheet, row, headers, col("portal"), _detect_portal(str(link)))
        _set_cell_by_header(offers_sheet, row, headers, col("last_action"), "Zaktualizowano")

        changed = previous_availability != availability_result.availability
        history_result = (
            "Niedostępna"
            if availability_result.availability == "Zamknięta"
            else "Zaktualizowano"
        )
        results.append(
            AvailabilityRefreshRow(
                offer_id=str(offer_id),
                company=str(company),
                title=str(title),
                link=str(link),
                previous_availability=str(previous_availability),
                availability=availability_result.availability,
                note=availability_result.note,
                changed=changed,
            )
        )
        _append_history_row(
            history_sheet,
            HistoryRecord(
                checked_at=today,
                offer_id=str(offer_id),
                company=str(company),
                title=str(title),
                link=str(link),
                result=history_result,
                checked_scope="Automatyczne sprawdzenie dostępności",
                note=(
                    f"Poprzednia dostępność: {previous_availability}. "
                    f"Obecna dostępność: {availability_result.availability}. "
                    f"{availability_result.note}"
                ),
                source=availability_result.final_url or str(link),
            ),
        )

    _refresh_dashboard_actions(dashboard_sheet, offers_sheet)
    _resize_known_table(offers_sheet)
    _resize_known_table(history_sheet)
    workbook.save(workbook_path)

    return AvailabilityRefreshSummary(
        checked_count=len(results),
        available_count=sum(result.availability == "Dostępna" for result in results),
        closed_count=sum(result.availability == "Zamknięta" for result in results),
        uncertain_count=sum(result.availability == "Niepewna" for result in results),
        changed_count=sum(result.changed for result in results),
        results=results,
    )


def refresh_offer_salaries_in_workbook(workbook_path: Path) -> SalaryRefreshSummary:
    from job_offer_analyzer.fetch_offer import fetch_offer_from_url

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")
    _ensure_workbook_writable(workbook_path)

    workbook = load_workbook(workbook_path)
    _normalize_workbook_sheets(workbook)
    offers_sheet = _get_sheet(workbook, OFFERS_SHEET)
    history_sheet = _get_sheet(workbook, HISTORY_SHEET)
    dashboard_sheet = _get_sheet(workbook, DASHBOARD_SHEET)
    _normalize_offer_headers(offers_sheet)
    _normalize_sheet_headers(history_sheet, HISTORY_HEADER_RENAMES)
    _ensure_sheet_headers(offers_sheet, SALARY_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_TRACKING_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_MATCH_HEADERS)
    _normalize_sheet_values(offers_sheet)
    _normalize_sheet_values(history_sheet)
    _normalize_history_actions(history_sheet)
    _normalize_dashboard_sheet(dashboard_sheet)
    _backfill_offer_tracking_columns(offers_sheet, history_sheet=history_sheet)

    headers = _header_positions(offers_sheet)
    results: list[SalaryRefreshRow] = []
    today = date.today()

    for row in range(2, offers_sheet.max_row + 1):
        offer_id = _cell_by_header(offers_sheet, row, headers, col("offer_id"))
        link = _cell_by_header(offers_sheet, row, headers, col("url"))
        if not offer_id or not link:
            continue

        company = _cell_by_header(offers_sheet, row, headers, col("company")) or ""
        title = _cell_by_header(offers_sheet, row, headers, col("job_title")) or ""

        try:
            draft = fetch_offer_from_url(str(link))
        except Exception as exc:
            results.append(
                SalaryRefreshRow(
                    offer_id=str(offer_id),
                    company=str(company),
                    title=str(title),
                    link=str(link),
                    salary_display=UNKNOWN_VALUE,
                    updated=False,
                    note=f"Nie udało się pobrać oferty: {exc}",
                )
            )
            continue

        salary = draft.salary
        if not _has_salary_values(salary):
            results.append(
                SalaryRefreshRow(
                    offer_id=str(offer_id),
                    company=str(company),
                    title=str(title),
                    link=str(link),
                    salary_display=UNKNOWN_VALUE,
                    updated=False,
                    note="Nie znaleziono widełek płacowych w danych oferty.",
                )
            )
            continue

        values = {col("rate_expectations"): salary.display_value}
        values.update(_salary_values(salary))
        _write_values(offers_sheet, row, headers, values)
        _set_cell_by_header(offers_sheet, row, headers, col("portal"), _detect_portal(str(link)))
        _set_cell_by_header(offers_sheet, row, headers, col("last_action"), "Zaktualizowano")
        _append_history_row(
            history_sheet,
            HistoryRecord(
                checked_at=today,
                offer_id=str(offer_id),
                company=str(company),
                title=str(title),
                link=str(link),
                result="Zaktualizowano",
                checked_scope="Aktualizacja stawek",
                note=f"Uzupełniono stawki z linku oferty: {salary.display_value}.",
                source=str(link),
            ),
        )

        results.append(
            SalaryRefreshRow(
                offer_id=str(offer_id),
                company=str(company),
                title=str(title),
                link=str(link),
                salary_display=salary.display_value,
                updated=True,
                note="Uzupełniono stawki z linku oferty.",
            )
        )

    _refresh_dashboard_actions(dashboard_sheet, offers_sheet)
    _resize_known_table(offers_sheet)
    _resize_known_table(history_sheet)
    workbook.save(workbook_path)

    return SalaryRefreshSummary(
        checked_count=len(results),
        updated_count=sum(result.updated for result in results),
        missing_count=sum(
            not result.updated and "Nie znaleziono" in result.note for result in results
        ),
        failed_count=sum(
            not result.updated and "Nie udało się" in result.note for result in results
        ),
        results=results,
    )


def refresh_cv_matches_in_workbook(
    workbook_path: Path, profile_path: Path, selected_link: str | None = None
) -> CvMatchRefreshSummary:
    from job_offer_analyzer.fetch_offer import fetch_offer_from_url
    from services.cv_matcher import (
        load_candidate_profile,
        match_offer_to_profile,
    )

    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")
    _ensure_workbook_writable(workbook_path)

    profile = load_candidate_profile(Path(profile_path))
    workbook = load_workbook(workbook_path)
    _normalize_workbook_sheets(workbook)
    offers_sheet = _get_sheet(workbook, OFFERS_SHEET)
    cv_match_sheet = _get_sheet(workbook, CV_MATCH_SHEET)
    dashboard_sheet = _get_sheet(workbook, DASHBOARD_SHEET)

    _normalize_offer_headers(offers_sheet)
    _normalize_sheet_headers(cv_match_sheet, CV_MATCH_HEADER_RENAMES)
    _ensure_sheet_headers(offers_sheet, SALARY_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_TRACKING_HEADERS)
    _ensure_sheet_headers(offers_sheet, OFFER_MATCH_HEADERS)
    _ensure_sheet_headers(cv_match_sheet, CV_MATCH_HEADERS)
    _normalize_sheet_values(offers_sheet)
    _normalize_dashboard_sheet(dashboard_sheet)
    _backfill_offer_tracking_columns(offers_sheet, cv_match_sheet)

    offers_headers = _header_positions(offers_sheet)
    cv_headers = _header_positions(cv_match_sheet)
    results: list[CvMatchRefreshRow] = []
    normalized_selected_link = selected_link.strip() if selected_link else ""
    matched_selected_link = False

    for row in range(2, offers_sheet.max_row + 1):
        offer_id = _cell_by_header(offers_sheet, row, offers_headers, col("offer_id"))
        link = _cell_by_header(offers_sheet, row, offers_headers, col("url"))
        if not offer_id or not link:
            continue

        link_text = str(link).strip()
        if normalized_selected_link:
            if link_text != normalized_selected_link:
                continue
            matched_selected_link = True
        elif _has_cv_match_value(
            _cell_by_header(offers_sheet, row, offers_headers, col("match_score"))
        ) or _has_cv_match_value(
            _cell_by_header(offers_sheet, row, offers_headers, col("cv_match"))
        ):
            continue

        company = str(_cell_by_header(offers_sheet, row, offers_headers, col("company")) or "")
        title = str(_cell_by_header(offers_sheet, row, offers_headers, col("job_title")) or "")
        fallback_text = _offer_text_from_row(offers_sheet, row, offers_headers)
        fetch_note = ""

        try:
            draft = fetch_offer_from_url(link_text)
            offer_text = " ".join(
                value
                for value in [
                    draft.title,
                    draft.company,
                    draft.must_have_summary,
                    draft.nice_to_have_summary,
                    draft.risks_notes,
                    draft.source_text,
                ]
                if value and value != UNKNOWN_VALUE
            )
        except Exception as exc:
            offer_text = fallback_text
            fetch_note = f"Nie pobrano strony, użyto danych z arkusza: {exc}"

        match_result = match_offer_to_profile(offer_text, profile)
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("match_score"),
            match_result.match_score,
        )
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("priority_code"),
            _priority_code(match_result.priority),
        )
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("technologies"),
            _technologies_from_match_result(match_result),
        )
        _set_cell_by_header(offers_sheet, row, offers_headers, col("portal"), _detect_portal(link_text))
        _set_cell_by_header(offers_sheet, row, offers_headers, col("last_action"), "Zaktualizowano")
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("cv_match"),
            f"{match_result.match_score}%",
        )
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("priority"),
            _priority_label(match_result.priority),
        )
        _set_cell_by_header(
            offers_sheet,
            row,
            offers_headers,
            col("missing_skills"),
            _missing_skills_text(match_result),
        )

        _clear_auto_cv_match_rows(cv_match_sheet, cv_headers, str(offer_id))
        _append_cv_match_rows(
            cv_match_sheet,
            cv_headers,
            offer_id=str(offer_id),
            company=company,
            title=title,
            link=link_text,
            match_result=match_result,
            fetch_note=fetch_note,
        )

        results.append(
            CvMatchRefreshRow(
                offer_id=str(offer_id),
                company=company,
                title=title,
                link=link_text,
                match_score=match_result.match_score,
                priority=match_result.priority,
                matched_skills=match_result.matched_skills,
                missing_skills=match_result.missing_skills,
                updated=True,
                note=match_result.short_reason if not fetch_note else f"{match_result.short_reason} {fetch_note}",
            )
        )

    if normalized_selected_link and not matched_selected_link:
        results.append(
            CvMatchRefreshRow(
                offer_id="",
                company="",
                title="",
                link=normalized_selected_link,
                match_score=0,
                priority="LOW",
                matched_skills=[],
                missing_skills=[],
                updated=False,
                note="Nie znaleziono zapisanej oferty z podanym linkiem.",
            )
        )

    _refresh_dashboard_actions(dashboard_sheet, offers_sheet)
    _resize_known_table(offers_sheet)
    _resize_known_table(cv_match_sheet)
    workbook.save(workbook_path)

    return CvMatchRefreshSummary(
        checked_count=len(results),
        updated_count=sum(result.updated for result in results),
        skipped_count=0,
        failed_count=sum(not result.updated for result in results),
        results=results,
    )


def _append_offer_row(sheet: Worksheet, offer_id: str, offer: OfferRecord) -> int:
    headers = _header_positions(sheet)
    row = _first_empty_row(sheet)
    _copy_row_style(sheet, source_row=2, target_row=row)

    source = offer.source or offer.link
    values = {
        col("offer_id"): offer_id,
        col("date_added"): offer.added_at,
        col("company"): offer.company,
        col("job_title"): offer.title,
        col("url"): offer.link,
        col("last_checked"): offer.last_checked_at,
        col("availability"): offer.availability,
        col("status"): offer.status,
        col("cv_match"): offer.cv_match,
        col("priority"): _priority_label(offer.priority),
        col("match_score"): _parse_match_score(offer.cv_match),
        col("priority_code"): _priority_code(offer.priority),
        col("technologies"): _technology_text(offer.technologies),
        col("portal"): _detect_portal(offer.link),
        col("last_action"): "Dodano",
        col("work_mode"): offer.work_mode,
        col("location"): offer.location,
        col("contract_type"): offer.contract_type,
        col("rate_expectations"): offer.rate_expectations,
        col("seniority"): offer.seniority,
        col("days_since_check"): offer.days_since_check,
        col("must_have_summary"): offer.must_have_summary,
        col("nice_to_have_summary"): offer.nice_to_have_summary,
        col("risks_notes"): offer.risks_notes,
        col("next_step"): offer.next_step,
        col("source"): source,
        col("missing_skills"): UNKNOWN_VALUE,
    }
    values.update(_salary_values(offer.salary))
    _write_values(sheet, row, headers, values)
    return row


def _append_history_row(sheet: Worksheet, history: HistoryRecord) -> int:
    headers = _header_positions(sheet)
    row = _first_empty_row(sheet)
    _copy_row_style(sheet, source_row=2, target_row=row)

    values = {
        col("checked_at"): history.checked_at,
        col("offer_ref_id"): history.offer_id,
        col("company"): history.company,
        col("job_title"): history.title,
        col("url"): history.link,
        col("result"): history.result,
        col("checked_scope"): history.checked_scope,
        col("note"): history.note,
        col("source"): history.source,
    }
    _write_values(sheet, row, headers, values)
    return row


def _append_question_placeholder_row(
    sheet: Worksheet, offer_id: str, offer: OfferRecord
) -> int:
    headers = _header_positions(sheet)
    if _question_row_exists(sheet, headers, offer_id):
        return 0

    row = _first_empty_row(sheet)
    _copy_row_style(sheet, source_row=2, target_row=row)

    values = {
        col("offer_ref_id"): offer_id,
        col("company"): offer.company,
        col("job_title"): offer.title,
        col("application_question"): "Nie odczytano pytań z formularza aplikacyjnego",
        col("required"): "Do sprawdzenia",
        col("answer_draft"): None,
        col("answer_status"): "Do sprawdzenia",
        col("comments"): (
            "Uzupełnić po wejściu w formularz aplikacyjny. "
            "Automatyczne pobranie opisu oferty nie zawiera pytań formularza."
        ),
    }
    _write_values(sheet, row, headers, values)
    return row


def _question_row_exists(
    sheet: Worksheet, headers: dict[str, int], offer_id: str
) -> bool:
    id_column = _column_for_header(headers, col("offer_ref_id"))
    if id_column is None:
        return False

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=id_column).value == offer_id:
            return True
    return False


def _offer_text_from_row(
    sheet: Worksheet, row: int, headers: dict[str, int]
) -> str:
    text_headers = [
        col("company"),
        col("job_title"),
        col("technologies"),
        col("work_mode"),
        col("location"),
        col("contract_type"),
        col("seniority"),
        col("must_have_summary"),
        col("nice_to_have_summary"),
        col("risks_notes"),
        col("rate_expectations"),
    ]
    values = [
        str(value)
        for header in text_headers
        if (value := _cell_by_header(sheet, row, headers, header))
        and str(value) != UNKNOWN_VALUE
    ]
    return " ".join(values)


def _has_cv_match_value(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return text not in {"", "TBD", UNKNOWN_VALUE}


def _missing_skills_text(match_result) -> str:
    if not match_result.missing_skills:
        return "Brak wykrytych braków"
    return "; ".join(match_result.missing_skills)


def _clear_auto_cv_match_rows(
    sheet: Worksheet, headers: dict[str, int], offer_id: str
) -> None:
    id_column = _column_for_header(headers, col("offer_ref_id"))
    category_column = _column_for_header(headers, col("category"))
    if id_column is None or category_column is None:
        return

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=id_column).value != offer_id:
            continue
        if sheet.cell(row=row, column=category_column).value != AUTO_MATCH_CATEGORY:
            continue
        for column in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=column).value = None


def _append_cv_match_rows(
    sheet: Worksheet,
    headers: dict[str, int],
    offer_id: str,
    company: str,
    title: str,
    link: str,
    match_result,
    fetch_note: str,
) -> None:
    for requirement in match_result.requirements:
        row = _first_empty_row(sheet)
        _copy_row_style(sheet, source_row=2, target_row=row)
        comment = requirement.comment
        if fetch_note:
            comment = f"{comment} {fetch_note}"

        values = {
            col("offer_ref_id"): offer_id,
            col("company"): company,
            col("job_title"): title,
            col("url"): link,
            col("category"): AUTO_MATCH_CATEGORY,
            col("offer_requirement"): requirement.requirement,
            col("in_profile"): "Tak" if requirement.has_skill else "Nie",
            col("evidence_skill"): requirement.evidence,
            col("match_strength"): f"{match_result.match_score}%",
            col("missing_to_learn"): requirement.missing_skill,
            col("importance"): _priority_label(match_result.priority),
            col("comment"): comment,
        }
        _write_values(sheet, row, headers, values)


def _history_from_offer(offer_id: str, offer: OfferRecord) -> HistoryRecord:
    return HistoryRecord(
        checked_at=offer.last_checked_at,
        offer_id=offer_id,
        company=offer.company,
        title=offer.title,
        link=offer.link,
        result="Dodano",
        checked_scope="Dodanie oferty",
        note="Dodano ofertę do arkusza Oferty.",
        source=offer.source or offer.link,
    )


def _history_duplicate_offer(offer_id: str, offer: OfferRecord) -> HistoryRecord:
    return HistoryRecord(
        checked_at=date.today(),
        offer_id=offer_id,
        company=offer.company,
        title=offer.title,
        link=offer.link,
        result="Pominięta jako duplikat",
        checked_scope="Dodanie oferty",
        note="Nie dodano nowego rekordu, ponieważ ten link już istnieje w arkuszu Oferty.",
        source=offer.source or offer.link,
    )


def _normalize_workbook_sheets(workbook) -> None:
    if CV_MATCH_SHEET in workbook.sheetnames:
        return
    if LEGACY_CV_MATCH_SHEET in workbook.sheetnames:
        workbook[LEGACY_CV_MATCH_SHEET].title = CV_MATCH_SHEET


def _get_sheet(workbook, sheet_name: str) -> Worksheet:
    wanted_names = [sheet_name, *SHEET_ALIASES.get(sheet_name, [])]
    wanted_keys = {_text_key(name) for name in wanted_names}
    for existing_name in workbook.sheetnames:
        if _text_key(existing_name) in wanted_keys:
            return workbook[existing_name]

    raise KeyError(f"Nie znaleziono arkusza: {sheet_name}")


def _header_positions(sheet: Worksheet) -> dict[str, int]:
    headers: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=column).value
        if value:
            header = str(value).strip()
            headers[header] = column
            headers[_text_key(header)] = column
    return headers


def _normalize_offer_headers(sheet: Worksheet) -> None:
    if _has_exact_header(sheet, col("priority_code_legacy")):
        target_header = (
            col("priority_code")
            if _has_exact_header(sheet, col("priority"))
            else col("priority")
        )
        _rename_exact_header(sheet, col("priority_code_legacy"), target_header)

    _normalize_sheet_headers(sheet, OFFER_HEADER_RENAMES)


def _has_exact_header(sheet: Worksheet, header: str) -> bool:
    return any(
        sheet.cell(row=1, column=column).value == header
        for column in range(1, sheet.max_column + 1)
    )


def _rename_exact_header(sheet: Worksheet, source_header: str, target_header: str) -> None:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=column).value == source_header:
            sheet.cell(row=1, column=column, value=target_header)


def _normalize_sheet_headers(sheet: Worksheet, renames: dict[str, str]) -> None:
    for column in range(1, sheet.max_column + 1):
        value = sheet.cell(row=1, column=column).value
        if isinstance(value, str) and value in renames:
            sheet.cell(row=1, column=column, value=renames[value])


def _ensure_sheet_headers(sheet: Worksheet, headers: list[str]) -> None:
    existing_headers = _header_positions(sheet)
    for header in headers:
        if _column_for_header(existing_headers, header) is not None:
            continue

        target_column = sheet.max_column + 1
        source_cell = sheet.cell(row=1, column=target_column - 1)
        target_cell = sheet.cell(row=1, column=target_column, value=header)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
        for row in range(2, sheet.max_row + 1):
            source_body_cell = sheet.cell(row=row, column=target_column - 1)
            target_body_cell = sheet.cell(row=row, column=target_column)
            if not source_body_cell.has_style:
                continue
            target_body_cell.font = copy(source_body_cell.font)
            target_body_cell.fill = copy(source_body_cell.fill)
            target_body_cell.border = copy(source_body_cell.border)
            target_body_cell.alignment = copy(source_body_cell.alignment)
            target_body_cell.number_format = source_body_cell.number_format
            target_body_cell.protection = copy(source_body_cell.protection)
        existing_headers[header] = target_column
        existing_headers[_text_key(header)] = target_column


def _backfill_offer_tracking_columns(
    offers_sheet: Worksheet,
    cv_match_sheet: Worksheet | None = None,
    history_sheet: Worksheet | None = None,
) -> None:
    headers = _header_positions(offers_sheet)
    technologies_by_offer = (
        _cv_technologies_by_offer(cv_match_sheet) if cv_match_sheet is not None else {}
    )
    latest_actions = (
        _latest_history_actions_by_offer(history_sheet)
        if history_sheet is not None
        else {}
    )

    for row in range(2, offers_sheet.max_row + 1):
        offer_id = _cell_by_header(offers_sheet, row, headers, col("offer_id"))
        link = _cell_by_header(offers_sheet, row, headers, col("url"))
        if not offer_id and not link:
            continue

        current_score = _cell_by_header(offers_sheet, row, headers, col("match_score"))
        parsed_current_score = _parse_match_score(current_score)
        legacy_score = _parse_match_score(
            _cell_by_header(offers_sheet, row, headers, col("cv_match"))
        )
        if parsed_current_score is not None:
            _set_cell_by_header(
                offers_sheet, row, headers, col("match_score"), parsed_current_score
            )
        elif legacy_score is not None:
            _set_cell_by_header(offers_sheet, row, headers, col("match_score"), legacy_score)

        current_priority_code = _cell_by_header(
            offers_sheet, row, headers, col("priority_code")
        )
        visible_priority = _cell_by_header(offers_sheet, row, headers, col("priority"))
        priority_code = _priority_code(current_priority_code) or _priority_code(
            visible_priority
        )
        priority_label = _priority_label(visible_priority) or _priority_label(
            priority_code
        )
        if priority_code:
            _set_cell_by_header(
                offers_sheet, row, headers, col("priority_code"), priority_code
            )
        if priority_label:
            _set_cell_by_header(
                offers_sheet, row, headers, col("priority"), priority_label
            )

        if link:
            _set_cell_by_header(offers_sheet, row, headers, col("portal"), _detect_portal(str(link)))

        current_action = _cell_by_header(offers_sheet, row, headers, col("last_action"))
        latest_action = latest_actions.get(str(offer_id)) if offer_id else None
        if _is_empty_value(current_action):
            _set_cell_by_header(
                offers_sheet, row, headers, col("last_action"), latest_action or "Dodano"
            )
        elif current_action == "Dodano" and latest_action not in {None, "Dodano"}:
            _set_cell_by_header(
                offers_sheet, row, headers, col("last_action"), latest_action
            )

        current_technologies = _cell_by_header(offers_sheet, row, headers, col("technologies"))
        normalized_technologies = _technology_text(current_technologies)
        if normalized_technologies:
            _set_cell_by_header(
                offers_sheet, row, headers, col("technologies"), normalized_technologies
            )
        elif _is_empty_value(current_technologies) and offer_id:
            technologies = technologies_by_offer.get(str(offer_id), "")
            if technologies:
                _set_cell_by_header(offers_sheet, row, headers, col("technologies"), technologies)


def _cv_technologies_by_offer(sheet: Worksheet) -> dict[str, str]:
    headers = _header_positions(sheet)
    technologies: dict[str, list[str]] = {}

    for row in range(2, sheet.max_row + 1):
        offer_id = _cell_by_header(sheet, row, headers, col("offer_ref_id"))
        if not offer_id:
            continue

        values = [
            _cell_by_header(sheet, row, headers, col("offer_requirement")),
            _cell_by_header(sheet, row, headers, col("evidence_skill")),
            _cell_by_header(sheet, row, headers, col("missing_to_learn")),
        ]
        for value in values:
            for technology in _split_technology_values(value):
                technologies.setdefault(str(offer_id), []).append(technology)

    return {
        offer_id: "; ".join(_dedupe_case_insensitive(values))
        for offer_id, values in technologies.items()
    }


def _latest_history_actions_by_offer(sheet: Worksheet) -> dict[str, str]:
    headers = _header_positions(sheet)
    id_column = _column_for_header(headers, col("offer_ref_id"))
    result_column = _column_for_header(headers, col("result"))
    if id_column is None or result_column is None:
        return {}

    actions: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        offer_id = sheet.cell(row=row, column=id_column).value
        result = sheet.cell(row=row, column=result_column).value
        if not offer_id or not result:
            continue
        action = _history_action_from_result(result, None)
        if action:
            actions[str(offer_id)] = action
    return actions


def _resize_known_table(sheet: Worksheet) -> None:
    table_name = None
    sheet_key = _text_key(sheet.title)
    for known_sheet_name, known_table_name in TABLE_NAMES.items():
        if _text_key(known_sheet_name) == sheet_key:
            table_name = known_table_name
            break

    if not table_name or table_name not in sheet.tables:
        return

    last_column = get_column_letter(sheet.max_column)
    last_row = max(sheet.max_row, 200)
    sheet.tables[table_name].ref = f"A1:{last_column}{last_row}"


def _normalize_sheet_values(sheet: Worksheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value in VALUE_RENAMES:
                cell.value = VALUE_RENAMES[cell.value]


def _normalize_history_actions(sheet: Worksheet) -> None:
    headers = _header_positions(sheet)
    result_column = _column_for_header(headers, col("result"))
    scope_column = _column_for_header(headers, col("checked_scope"))
    note_column = _column_for_header(headers, col("note"))
    if result_column is None:
        return

    for row in range(2, sheet.max_row + 1):
        result = sheet.cell(row=row, column=result_column).value
        scope = (
            sheet.cell(row=row, column=scope_column).value
            if scope_column is not None
            else None
        )
        normalized_result = _history_action_from_result(result, scope)
        if not normalized_result:
            continue

        if result != normalized_result:
            sheet.cell(row=row, column=result_column, value=normalized_result)
            if note_column is not None and result:
                note = sheet.cell(row=row, column=note_column).value
                sheet.cell(
                    row=row,
                    column=note_column,
                    value=_append_migration_note(note, f"Poprzedni wynik: {result}."),
                )

        if scope == "Ręczny wpis testowy" and scope_column is not None:
            sheet.cell(row=row, column=scope_column, value="Dodanie oferty")


def _history_action_from_result(result: object, scope: object) -> str | None:
    if _is_empty_value(result):
        return None

    text = str(result).strip()
    if text in {"Dodano", "Zaktualizowano", "Niedostępna", "Pominięta jako duplikat"}:
        return text
    if scope == "Ręczny wpis testowy":
        return "Dodano"
    if text == "Zamknięta":
        return "Niedostępna"
    if text in {"Dostępna", "Niepewna"}:
        return "Zaktualizowano"
    return None


def _append_migration_note(note: object, migration_note: str) -> str:
    if not note:
        return migration_note
    note_text = str(note)
    if migration_note in note_text:
        return note_text
    return f"{note_text} {migration_note}"


def _normalize_dashboard_sheet(sheet: Worksheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value in DASHBOARD_TEXT_RENAMES:
                cell.value = DASHBOARD_TEXT_RENAMES[cell.value]

    sheet["B4"] = '=COUNTA(Oferty!A2:A200)'
    sheet["B5"] = '=COUNTIF(Oferty!G2:G200,"Dostępna")'
    sheet["B6"] = '=COUNTIF(Oferty!H2:H200,"Do analizy")'
    sheet["B7"] = '=COUNTIF(Oferty!H2:H200,"Aplikowano")'
    sheet["B8"] = '=COUNTIF(Oferty!J2:J200,"Wysoki")'
    sheet["B9"] = '=COUNTIF(Oferty!P2:P200,">14")'


def _refresh_dashboard_actions(dashboard_sheet: Worksheet, offers_sheet: Worksheet) -> None:
    _refresh_dashboard_formulas(dashboard_sheet, offers_sheet)
    action_rows = _dashboard_action_rows(offers_sheet)
    last_action_row = max(dashboard_sheet.max_row, 4 + len(action_rows), 9)

    for row in range(5, last_action_row + 1):
        for column in range(4, 9):
            dashboard_sheet.cell(row=row, column=column).value = None

    for offset, action in enumerate(action_rows, start=5):
        _copy_dashboard_action_style(dashboard_sheet, target_row=offset)
        for column_offset, value in enumerate(action, start=4):
            dashboard_sheet.cell(row=offset, column=column_offset, value=value)


def _dashboard_action_rows(offers_sheet: Worksheet) -> list[tuple[object, ...]]:
    headers = _header_positions(offers_sheet)
    actions: list[tuple[int, tuple[object, ...]]] = []

    for row in range(2, offers_sheet.max_row + 1):
        offer_id = _cell_by_header(offers_sheet, row, headers, col("offer_id"))
        if not offer_id:
            continue

        status = _cell_by_header(offers_sheet, row, headers, col("status"))
        availability = _cell_by_header(offers_sheet, row, headers, col("availability"))
        next_step = _cell_by_header(offers_sheet, row, headers, col("next_step"))
        if status in ACTIVE_DASHBOARD_EXCLUDED_STATUSES:
            continue
        if availability in ACTIVE_DASHBOARD_EXCLUDED_AVAILABILITY:
            continue
        if not next_step or next_step == UNKNOWN_VALUE:
            continue

        actions.append(
            (
                _offer_id_sort_key(str(offer_id)),
                (
                    offer_id,
                    _cell_by_header(offers_sheet, row, headers, col("company")),
                    _cell_by_header(offers_sheet, row, headers, col("job_title")),
                    status,
                    next_step,
                ),
            )
        )

    actions.sort(key=lambda item: item[0], reverse=True)
    return [action for _, action in actions]


def _refresh_dashboard_formulas(
    dashboard_sheet: Worksheet, offers_sheet: Worksheet
) -> None:
    headers = _header_positions(offers_sheet)
    id_column = _column_letter_for_header(headers, col("offer_id"))
    availability_column = _column_letter_for_header(headers, col("availability"))
    status_column = _column_letter_for_header(headers, col("status"))
    priority_column = _column_letter_for_header(headers, col("priority"))
    days_column = _column_letter_for_header(headers, col("days_since_check"))

    if id_column:
        dashboard_sheet["B4"] = f"=COUNTA(Oferty!{id_column}2:{id_column}200)"
    if availability_column:
        dashboard_sheet["B5"] = (
            f'=COUNTIF(Oferty!{availability_column}2:{availability_column}200,'
            '"Dostępna")'
        )
    if status_column:
        dashboard_sheet["B6"] = (
            f'=COUNTIF(Oferty!{status_column}2:{status_column}200,"Do analizy")'
        )
        dashboard_sheet["B7"] = (
            f'=COUNTIF(Oferty!{status_column}2:{status_column}200,"Aplikowano")'
        )
    if priority_column:
        dashboard_sheet["B8"] = (
            f'=COUNTIF(Oferty!{priority_column}2:{priority_column}200,"Wysoki")'
        )
    if days_column:
        dashboard_sheet["B9"] = f'=COUNTIF(Oferty!{days_column}2:{days_column}200,">14")'


def _cell_by_header(
    sheet: Worksheet, row: int, headers: dict[str, int], header: str
) -> object:
    column = _column_for_header(headers, header)
    if column is None:
        return None
    return sheet.cell(row=row, column=column).value


def _set_cell_by_header(
    sheet: Worksheet,
    row: int,
    headers: dict[str, int],
    header: str,
    value: object,
) -> None:
    column = _column_for_header(headers, header)
    if column is None:
        raise ValueError(f"Brak kolumny w arkuszu {sheet.title}: {header}")
    sheet.cell(row=row, column=column, value=value)


def _offer_id_sort_key(offer_id: str) -> int:
    match = OFFER_ID_PATTERN.match(offer_id.strip())
    if not match:
        return 0
    return int(match.group(1))


def _copy_dashboard_action_style(sheet: Worksheet, target_row: int) -> None:
    source_row = 5
    if target_row == source_row:
        return

    for column in range(4, 9):
        source_cell = sheet.cell(row=source_row, column=column)
        target_cell = sheet.cell(row=target_row, column=column)
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def _write_values(
    sheet: Worksheet, row: int, headers: dict[str, int], values: dict[str, object]
) -> None:
    missing_headers = [
        header for header in values if _column_for_header(headers, header) is None
    ]
    if missing_headers:
        joined_headers = ", ".join(missing_headers)
        raise ValueError(f"Brak kolumn w arkuszu {sheet.title}: {joined_headers}")

    for header, value in values.items():
        column = _column_for_header(headers, header)
        sheet.cell(row=row, column=column, value=value)


def _column_for_header(headers: dict[str, int], header: str) -> int | None:
    direct_column = headers.get(header) or headers.get(_text_key(header))
    if direct_column is not None:
        return direct_column

    for alias in HEADER_ALIASES.get(header, []):
        alias_column = headers.get(alias) or headers.get(_text_key(alias))
        if alias_column is not None:
            return alias_column

    return None


def _column_letter_for_header(headers: dict[str, int], header: str) -> str | None:
    column = _column_for_header(headers, header)
    if column is None:
        return None
    return get_column_letter(column)


def _find_offer_id_by_link(sheet: Worksheet, link: str) -> str | None:
    headers = _header_positions(sheet)
    id_column = _column_for_header(headers, col("offer_id"))
    link_column = _column_for_header(headers, col("url"))
    if id_column is None or link_column is None:
        return None

    normalized_link = _normalize_link(link)
    if not normalized_link:
        return None

    for row in range(2, sheet.max_row + 1):
        existing_link = sheet.cell(row=row, column=link_column).value
        if _normalize_link(existing_link) != normalized_link:
            continue

        offer_id = sheet.cell(row=row, column=id_column).value
        return str(offer_id) if offer_id else None

    return None


def _normalize_link(link: object) -> str:
    if link is None:
        return ""
    return str(link).strip().rstrip("/").lower()


def _detect_portal(url: str) -> str:
    try:
        parsed_url = urlparse(url if "://" in url else f"https://{url}")
    except ValueError:
        return "Nieznany"

    host = parsed_url.netloc.lower().removeprefix("www.")
    full_url = f"{host}{parsed_url.path}".lower()
    if "justjoin.it" in host or "justjoinit" in host:
        return "Just Join IT"
    if "nofluffjobs.com" in host:
        return "No Fluff Jobs"
    if "linkedin.com" in host:
        return "LinkedIn"
    if "testdevjobs" in full_url:
        return "TestDevJobs"
    return "Nieznany"


def _priority_code(value: object) -> str | None:
    if _is_empty_value(value):
        return None

    priority_map = {
        "high": "HIGH",
        "medium": "MEDIUM",
        "low": "LOW",
    }
    priority_map.update(PRIORITY_LABEL_TO_CODE)
    return priority_map.get(_text_key(str(value)))


def _priority_label(value: object) -> str | None:
    priority_code = _priority_code(value)
    if priority_code is None:
        return None
    return PRIORITY_CODE_TO_LABEL[priority_code]


def _normalize_priority(value: object) -> str | None:
    return _priority_code(value)


def _parse_match_score(value: object) -> int | None:
    if _is_empty_value(value):
        return None

    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))

    match = re.search(r"\d+(?:[,.]\d+)?", str(value))
    if not match:
        return None

    return int(round(float(match.group(0).replace(",", "."))))


def _technology_text(value: object) -> str | None:
    technologies = _split_technology_values(value)
    if not technologies:
        return None
    return "; ".join(_dedupe_case_insensitive(technologies))


def _technologies_from_match_result(match_result) -> str | None:
    technologies: list[str] = []
    for requirement in match_result.requirements:
        technologies.extend(
            _split_technology_values(requirement.evidence)
            + _split_technology_values(requirement.missing_skill)
            + _split_technology_values(requirement.requirement)
        )
    return _technology_text("; ".join(technologies))


def _split_technology_values(value: object) -> list[str]:
    if _is_empty_value(value):
        return []

    ignored_values = {
        "brak",
        "brak dopasowania w profilu",
        "do recznej weryfikacji",
        "do ręcznej weryfikacji",
        "nie dotyczy",
        "wymaga nauki lub dopisania do profilu",
    }
    values = re.split(r"[;\n|]+", str(value))
    cleaned_values = []
    for item in values:
        cleaned_item = re.sub(r"\s+", " ", item).strip(" .,")
        if not cleaned_item or _text_key(cleaned_item) in ignored_values:
            continue
        if cleaned_item in {"-", "_"}:
            continue
        recognized_technologies = _recognized_technologies(cleaned_item)
        if recognized_technologies:
            cleaned_values.extend(recognized_technologies)
            continue
        if len(cleaned_item) <= 40:
            cleaned_values.append(cleaned_item)
    return cleaned_values


def _recognized_technologies(text: str) -> list[str]:
    matches: list[str] = []
    for technology, aliases in TECHNOLOGY_ALIASES.items():
        if any(_contains_technology_alias(text, alias) for alias in aliases):
            matches.append(technology)
    return matches


def _contains_technology_alias(text: str, alias: str) -> bool:
    pattern = rf"(?<![a-z0-9#+]){re.escape(alias.lower())}(?![a-z0-9#+])"
    return re.search(pattern, text.lower()) is not None


def _dedupe_case_insensitive(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        key = _text_key(value)
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _is_empty_value(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text in {"", "TBD", UNKNOWN_VALUE, "Unknown", "Nieznany"}


def _salary_values(salary: SalaryInfo) -> dict[str, object]:
    return {
        col("salary_source"): salary.original_text,
        col("currency"): salary.currency,
        col("salary_min"): salary.amount_min,
        col("salary_max"): salary.amount_max,
        col("salary_period"): salary.period,
        col("tax_type"): salary.tax_type,
        col("exchange_rate"): salary.exchange_rate_to_pln,
        col("exchange_rate_date"): salary.exchange_rate_date,
        col("pln_min_monthly"): salary.pln_min_monthly,
        col("pln_max_monthly"): salary.pln_max_monthly,
        col("pln_min_hourly"): salary.pln_min_hourly,
        col("pln_max_hourly"): salary.pln_max_hourly,
        col("conversion_assumptions"): salary.conversion_assumptions,
    }


def _has_salary_values(salary: SalaryInfo) -> bool:
    return salary.currency != UNKNOWN_VALUE and salary.amount_min is not None


def _text_key(text: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", text.translate(POLISH_TRANSLATION))
    ascii_text = "".join(char for char in ascii_text if not unicodedata.combining(char))
    return ascii_text.strip().lower()


def _first_empty_row(sheet: Worksheet) -> int:
    for row in range(2, sheet.max_row + 2):
        values = [
            sheet.cell(row=row, column=column).value
            for column in range(1, sheet.max_column + 1)
        ]
        if all(value is None for value in values):
            return row
    return sheet.max_row + 1


def _next_offer_id(sheet: Worksheet) -> str:
    max_number = 0
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row=row, column=1).value
        if not isinstance(value, str):
            continue

        match = OFFER_ID_PATTERN.match(value.strip())
        if match:
            max_number = max(max_number, int(match.group(1)))

    return f"JOB-{max_number + 1:03d}"


def _copy_row_style(sheet: Worksheet, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return

    for column in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row, column=column)
        target_cell = sheet.cell(row=target_row, column=column)

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def _ensure_workbook_writable(workbook_path: Path) -> None:
    try:
        with open(workbook_path, "a+b"):
            pass
    except PermissionError as exc:
        raise PermissionError(
            "Nie można zapisać pliku Excel. Zamknij skoroszyt w Excelu i spróbuj ponownie."
        ) from exc
